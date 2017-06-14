#! python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 23 17:01:49 2017

The goal of this script is to send a mail with the vocabulary I must learn.
The goal is to learn five words per day.
If the script didn't run in the past days it should send the missing vocabulary.

@author: Arthur Baucour
"""

import openpyxl #  excel files
import datetime # date format


#==============================================================================
# Parameters

filename = 'Ewha study guide 1-1 - edit.xlsx'    # Filename
path = './Vocabulary lists (edited)/'       # Path to the file
dt_start = datetime.datetime(2017, 6, 9)    # Starting day: Year, month, day
words_per_day = 5                           # Words learned per day


#==============================================================================
# Open the file and get the data 

wb = openpyxl.load_workbook(path + filename) # Open the file we will use
sheet = wb.get_sheet_by_name('Sheet1') # Get the sheet


#==============================================================================
# Write datetime from the Starting day to the end of the list

delta = datetime.timedelta(days=1)

if dt_start.strftime('%m/%d/%Y') != sheet.cell(row=1, column=3).value: # If the start time is different
    print('Changing the dates')
    i=1
    while (sheet.cell(row=i, column=1).value): # While there is a value
        test = (i-1) // words_per_day
        sheet.cell(row=i, column=3).value = (dt_start + delta * test).strftime('%m/%d/%Y')
        i += 1


#==============================================================================
# List of words for the day

dt = datetime.datetime.now()

ls_today = []   # List of words and translation to learn for the day
i = 1           # Initialize counter
date_test = datetime.datetime(2000,1,1) # Initialize, don't really care

# List of missing vocabulary
while (date_test < dt):
    date_test = datetime.datetime.strptime(sheet.cell(row=i, column=3).value, '%m/%d/%Y') # Used to count
    if not (sheet.cell(row=i, column=4).value) and (date_test < dt): # if the row is not marked by an 'x'
        ls_today.append([sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value])
        sheet.cell(row=i, column=4).value = 'x' # Mark the word
    i += 1


wb.save(path + filename)


#==============================================================================
# Print vocabulary

for i in range(len(ls_today)):
    print(ls_today[i][0], ls_today[i][1])
       
    
#======================
#
#import smtplib
#
#smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
#smtpObj.ehlo()
#smtpObj.starttls()
#smtpObj.login('test@gmail.com', '')
#
#smtpObj.sendmail('test@gmail.com', 'test2@gmail.com', 'Subject: test\nTEST TEST')
#
#smtpObj.quit()